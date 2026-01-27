import xml.etree.ElementTree as ET
import csv
import gzip
import os
import glob
import datetime
import openpyxl 

def safe_find(element, path):
    found = element.find(path)
    return found.text if found is not None else ""

def parse_dmarc(xml_file):
    # Handle .gz compressed reports
    if xml_file.endswith(".gz"):
        with gzip.open(xml_file, "rb") as f:
            xml_content = f.read()
        root = ET.fromstring(xml_content)
    else:
        tree = ET.parse(xml_file)
        root = tree.getroot()

    records = []
    for record in root.findall("record"):
        row = record.find("row")
        identifiers = record.find("identifiers")
        auth_results = record.find("auth_results")

        records.append({
            "Source IP": safe_find(row, "source_ip"),
            "Count": safe_find(row, "count"),
            "Disposition": safe_find(row, "policy_evaluated/disposition"),
            "DKIM": safe_find(auth_results, "dkim/result"),
            "SPF": safe_find(auth_results, "spf/result"),
            "Envelope From": safe_find(identifiers, "envelope_from"),
            "Header From": safe_find(identifiers, "header_from"),
            "Envelope To": safe_find(identifiers, "envelope_to")
        })
    return records

from collections import defaultdict

def summarize_by_ip(records):
    summary = defaultdict(lambda: {"Total": 0, "Pass": 0, "Fail": 0})
    for rec in records:
        ip = rec["Source IP"]
        count = int(rec["Count"]) if rec["Count"].isdigit() else 1
        summary[ip]["Total"] += count

        # Simple pass/fail logic: if both DKIM and SPF == "pass"
        if rec["DKIM"].lower() == "pass" and rec["SPF"].lower() == "pass":
            summary[ip]["Pass"] += count
        else:
            summary[ip]["Fail"] += count

    # Convert to list of dicts for CSV export
    summary_records = []
    for ip, stats in summary.items():
        summary_records.append({
            "Source IP": ip,
            "Total Messages": stats["Total"],
            "Passes": stats["Pass"],
            "Fails": stats["Fail"],
            "Compliance %": round((stats["Pass"] / stats["Total"]) * 100, 2)
        })
    return summary_records

import datetime

def export_to_csv(records, filename=None):
    if not records:
        print("No records found.")
        return
    
    summaries_folder = os.path.join(os.path.dirname(__file__), "dmarc_summaries")
    os.makedirs(summaries_folder, exist_ok=True)

    if filename is None:
        timestamp = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
        filename = f"dmarc_summary_{timestamp}.csv"
    full_path = os.path.join(summaries_folder, filename)

    fieldnames = records[0].keys()
    with open(full_path, "w", newline="") as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(records)

    print(f"DMARC summary written to {full_path}")

from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule

def export_to_excel(raw_records, summary_records, filename=None):
    summaries_folder = os.path.join(os.path.dirname(__file__), "dmarc_summaries")
    os.makedirs(summaries_folder, exist_ok=True)

    if filename is None:
        timestamp = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
        filename = f"dmarc_report_{timestamp}.xlsx"

    full_path = os.path.join(summaries_folder, filename)

    wb = openpyxl.Workbook()

    # Raw Records sheet
    ws_raw = wb.active
    ws_raw.title = "Raw Records"
    if raw_records:
        headers = list(raw_records[0].keys())
        ws_raw.append(headers)
        for rec in raw_records:
            ws_raw.append([rec[h] for h in headers])

    # IP Summary sheet
    ws_summary = wb.create_sheet(title="IP Summary")
    if summary_records:
        headers = list(summary_records[0].keys())
        ws_summary.append(headers)
        for rec in summary_records:
            ws_summary.append([rec[h] for h in headers])

        # --- Conditional formatting on Compliance % column ---
        compliance_col = headers.index("Compliance %") + 1  # column index
        col_letter = openpyxl.utils.get_column_letter(compliance_col)
        data_range = f"{col_letter}2:{col_letter}{len(summary_records)+1}"

        # Green fill for >= 80%
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        ws_summary.conditional_formatting.add(
            data_range,
            CellIsRule(operator="greaterThanOrEqual", formula=["80"], fill=green_fill)
        )

        # Red fill for < 80%
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        ws_summary.conditional_formatting.add(
            data_range,
            CellIsRule(operator="lessThan", formula=["80"], fill=red_fill)
        )

    wb.save(full_path)
    print(f"DMARC Excel report written to {full_path}")


if __name__ == "__main__":
    try:
        reports_folder = os.path.join(os.path.dirname(__file__), "reports")
        all_records = []

        for xml_file in glob.glob(os.path.join(reports_folder, "*.gz")):
            print(f"Parsing {xml_file}...")
            all_records.extend(parse_dmarc(xml_file))

        if all_records:
            summary_records = summarize_by_ip(all_records)
            export_to_excel(all_records, summary_records)  # <-- new Excel export
        else:
            print("No DMARC .gz files found in reports folder.")

    except Exception as e:
        import traceback
        print("\n❌ An error occurred while running the DMARC parser:\n")
        traceback.print_exc()  # shows full error + line number

    # 👇 This always runs, success or error
    # input("\nPress Enter to return to the hub...")

